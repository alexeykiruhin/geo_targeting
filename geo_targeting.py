#Import
import re
import sys
import os
import geoUI # GUI главного окна
import dialog # Диалоговое окно

from PyQt5 import QtWidgets
from openpyxl import load_workbook

#Москва и Питер реализовать чекбокс с областями или без если с областями то
#москва 12 и 13 а питер 28 33, если без областей то 700 и 753



class Dialog(QtWidgets.QDialog, dialog.Ui_Dialog):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.btn.clicked.connect(self.choise)
        self.id1 = 0
        self.id2 = 0

    #подставляем названия в соотв. поля
    def setName(self, name1, name2, id1, id2):
        self.name_1.setText(name1)
        self.name_2.setText(name2)
        self.id1 = id1
        self.id2 = id2

    #обработка чекбокса - выбор варианта
    def choise(self):
        if self.name_1.checkState() == 2:
            Dialog.done(self, self.id1)
        elif self.name_2.checkState() == 2:
            Dialog.done(self, self.id2)

class App(QtWidgets.QMainWindow, geoUI.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.search.clicked.connect(self.search_id)
        self.clear.clicked.connect(self.clear_fields)
        self.all_cities = []
        self.list_cities = []
        self.list_rows = []
        self.not_found_list = []
        self.files = os.listdir(path=".")
        self.excel_file_name = 'geo.xlsx'
        #поиск эксель файла и его использование если нет geo.xlsx,
        #выбор файла или закрытие программы
        self.check_excel_file()
        try:
            self.wb = load_workbook(filename = self.excel_file_name)
        except FileNotFoundError as e:
            print('Файл не найден')
        #self.sheet = self.wb.get_sheet_by_name('Sheet1')
        self.sheet = self.wb.worksheets[0]
        #количество строк в екселе
        self.row_count = self.sheet.max_row
        print(self.row_count)
        self.excel_to_list()
        self.duble_words = 0

    def check_excel_file(self):
        try:
            i = self.files.index(self.excel_file_name)
        except ValueError:
            z = 0
            rash = []
            for f in self.files:
                z += 1
                rash = f.split('.')
                if len(rash) == 1:
                    continue
                if rash[1] == 'xlsx':
                    self.excel_file_name = rash[0] + '.' + rash[1]
                    print(self.excel_file_name)
                    print (z)

            #добавить условие если вообще нет файла

    #считываем эксель файл и записываем названия городов в список all_cities
    def excel_to_list(self):
        for i in range(2, self.row_count + 1):
            self.all_cities.append(self.get_cell(i-2, 2))

    #обработка введенных данных (названия городов) и запись в список list_cities
    def get_list_cities(self):
        # убираем лишние знаки, если список городов ввели через запятую
        #cities - название поля в GUI
        reg = " |,|\n"

        cit = self.cities.toPlainText().title()
        self.list_cities = re.split(reg, cit)

    #метод для проверки количества городов например Москва 2 шт
    #поиск области в дубликатах, по соседям перед этим городом и после,
    #проверяем на совподении по области 3й столбей в экселе
    #поиск дубликатов:
    #в списке введенных городов
    #в списке всех городов
    def duplicate_check(self, search):
        try:
            num_row = self.all_cities.index(search)
        except ValueError:
            #обработка события когда город последний в списке и он не нашёлся то в х будет ошибка
            #т.к. тут -  self.list_cities[self.list_cities.index(search) + 1] ничего не будет IndexError
            #print(self.list_cities.index(search) + 1)
            try:
                if self.list_cities[self.list_cities.index(search) + 1] == 'Область':
                    x = self.list_cities[self.list_cities.index(search)] + ' обл.'
                elif self.list_cities[self.list_cities.index(search) + 1] == 'Округ':
                    x = self.list_cities[self.list_cities.index(search)] + ' округ'
                elif self.list_cities[self.list_cities.index(search) + 1] == 'Край':
                    x = self.list_cities[self.list_cities.index(search)] + ' край'
                elif self.list_cities[self.list_cities.index(search) + 1] == 'Автономный'\
                 or self.list_cities[self.list_cities.index(search) + 1] == 'Автономная':
                    x = self.list_cities[self.list_cities.index(search)] + ' АО'
                    index = self.list_cities.index(search) + 2
                    self.list_cities.pop(index)
                elif self.list_cities[self.list_cities.index(search)] == 'Республика'\
                 or self.list_cities[self.list_cities.index(search)] == 'Респ':
                    x = 'Респ. ' + self.list_cities[self.list_cities.index(search) + 1]
                else:
                    x = self.list_cities[self.list_cities.index(search)] + ' '\
                     + self.list_cities[self.list_cities.index(search) + 1]
            except IndexError:
                if search == '':
                    return -1
                self.not_found_list.append(search)
                return -1

            #проверка есть ли вообще теперь такой город, если нет то записываем
            #его в отдельный список и когда найдутся все города выводим те которые не нашли
            #not_found_list - список ненайденных городов
            try:
                num_row = self.all_cities.index(x)
            except ValueError:
                if search == '' or search == 'Округ' or search == 'И' or search == 'Область':
                    return -1
                self.not_found_list.append(search)
                return -1
            else:
                self.duble_words += 1
                index = self.list_cities.index(search)
                self.list_cities.pop(index)
                self.list_cities.pop(index)
                self.list_cities.insert(index, x)
                num_row = self.all_cities.index(x)

        #else:
            #print('95 строка')


        #проверить есть ли в веденном списке городов дубликаты, если есть то выбираешь
        #первый потом второй
        #если в веденном списке нет дубликатов но они есть в общем, вызвать диалоговое
        #окно с выбором города( указав облассть и страну)
        #перед этим сравнив города до и после дубликата, если совпадений нет или
        #дубликат есть у первого города в списке то выводим диаоодиалоговое окнокно
        #с выбором города
        try:
            double_row = self.all_cities.index(search, num_row + 1)
            # сравнить названия дубликатов
        except ValueError:
            row = num_row + 2 #номер строки в ексель файле
            self.list_rows.append(row)
            return self.get_cell(num_row, 1) #возвращаем айди из экселя
        else: # если чекбокс активен то москва и питер с областями(активный чекбокс = 2)
            if self.get_cell(num_row, 2) != self.get_cell(double_row, 2):
                return self.get_cell(num_row, 1)
            else:
                if (self.region.checkState() == 2 and self.get_cell(num_row, 2) == 'Москва'):
                    return '12,13'
                elif (self.region.checkState() == 2 and self.get_cell(num_row, 2) == 'Санкт-Петербург'):
                    return '28,33'
                elif (self.region.checkState() == 0 and self.get_cell(num_row, 2) == 'Москва'):
                    return 700
                elif (self.region.checkState() == 0 and self.get_cell(num_row, 2) == 'Санкт-Петербург'):
                    return 756
                else:
                    #обработка названий из 2х слов нижний новгород
                    #return(num_row + 2, double_row + 2)#номера строк в ексель файле
                    #если 2 города нашлось, то делаем выбор по области
                    if len(self.list_rows) == 0:
                        try:
                            x = self.list_cities[self.list_cities.index(search) + 1]
                        except IndexError:
                            #переводим номер строки в название города и
                            #информацию о нем
                            #вызов дивлогового окна
                            return self.call_dialog(num_row, double_row)

                        #из id_to_name возвращается название и в last ищется первое совпадение
                        #print(x)
                        #last = self.all_cities.index(x)
                    elif self.list_rows[-1] != 0:
                        last = self.list_rows[-1] - 2# последний найденный город
                    #(-2 т.к в get_cell к строке добавляю 2)
                    if self.get_cell(last, 3) == self.get_cell(num_row, 3):
                        return self.get_cell(num_row, 1)
                    elif self.get_cell(last, 3) == self.get_cell(double_row, 3):
                        return self.get_cell(double_row, 1)
                    else:
                        #вызов дивлогового окна
                        return self.call_dialog(num_row, double_row)
                        #num_row - первый город - name_1
                        #double_row - второй город - name_2

    #вызов дивлогового окна
    def call_dialog(self, num_row, double_row):
        first = self.row_to_name(num_row)
        second = self.row_to_name(double_row)
        dialog = Dialog()
        dialog.setName(first, second, self.get_cell(num_row, 1), self.get_cell(double_row, 1))
        y = dialog.exec_()
        return y

    def get_cell(self, row, column):
        return (self.sheet.cell(row=row + 2, column=column).value)

    #метод поиска айди по списку городов
    def search_id(self):
        self.id.clear()
        self.not_found.clear()
        self.get_list_cities()
        self.not_found_list = []
        search_count = 0
        for i in self.list_cities: # cities = all_cities, search = list_cities
            id = self.duplicate_check(i) #возвращать флаг что город не найден и перейти к следующиму шагу
            if id == -1:
                continue
            if type(id) is str:
                self.id.insertPlainText(id + ',')
                search_count += 2
            else:
                self.id.insertPlainText(str(id) + ',')
                search_count += 1
        #список ненайденых городов
        self.not_found.insertPlainText(self.nf_list_to_str(self.not_found_list))
        #добавляем цифру найденных или ненайденных городов и их айди
        self.label.setText('Cities:' + ' ' + str(len(self.list_cities)))
        self.label_2.setText('Id:' + ' ' + str(search_count))
        self.label_3.setText('Сities not found:' + ' ' + str(len(self.not_found_list)))

    def nf_list_to_str(self, nf_list):
        not_found_str = ''
        for nf in nf_list:
            not_found_str += nf + ' '
        return not_found_str

    def clear_fields(self):
        self.cities.clear()
        self.id.clear()
        self.not_found.clear()

    def id_to_name(self, id):
        all_id = []
        #создаем список всех айди
        for i in range(2, self.row_count + 1):
            all_id.append(self.get_cell(i-2, 1))

        row = all_id.index(id)
        print(row)
        name = self.get_cell(row, 2)
        print(name)

        return name

    def name_to_row(self, name):
        row = self.all_cities.index(name)
        return row + 2

    def row_to_name(self, row):
        name_city = self.get_cell(row, 2)
        id_region = self.get_cell(row, 3)
        name_region = self.id_to_name(id_region)
        id_country = self.get_cell(self.name_to_row(name_region), 3)
        name_country = self.id_to_name(id_country)
        #склеиваем данные и возвращаем строку
        return name_city + ' ' + name_region + ' ' + name_country

def main():
    app = QtWidgets.QApplication(sys.argv)
    window = App()
    window.show()
    app.exec_()

if __name__ == '__main__':
    main()
